# imports
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re

#### SAMPLE FIGURES

# INPUTS & OUTPUTS: this code takes in uuids from an excel spreadsheet with the uuid being the first column in the table. first row is the csv file name,
# followed by 'uuid', and the actual data starts on row 3. it reads in the data from the excel sheet, then outputs as a new sheet
# to the same excel file it read in data from, but just to new sheets that are initialized below. the data sheet remains untouched during the export process.

# excel file name, no complete path needed. as long as it's in the same folder/location as the script
file_path = 'Figure Generation Samples.xlsx'

# name of the sheet in the excel file that has all the data in it
sheet_name = 'Data'

# putting all the data that script reads in into a data frame. this data frame is for the sample generation (colored figure)
# skiprows is in there because csv->excel = the first row of the excel file having the csv file name. get rid if not exporting from a csv file
df_main = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=1)

# resets the index after skipping the first row (title row)
df_main.reset_index(drop=True, inplace=True)

# sets max number of rows to none when you want to view the dataframe data on jupyter notebook
# pd.set_option('display.max_rows', None)

# extracting sample type (D.FLOW, TIS, etc.) from the data
def get_sample_type(uuid):
    return uuid.split('-')[0]

# extracting lab name (FLY, GRI, etc.) from the data
def get_lab(uuid):
    lab = uuid[-8:]
    return ''.join(re.findall('[A-Za-z]', lab))# Extract the last three letters after splitting

# making a 'sample type' and a 'lab' column in the dataframe to house the sample type and the lab
df_main['Sample Type'] = df_main['uuid'].apply(get_sample_type)
df_main['Lab'] = df_main['uuid'].apply(get_lab)

# dropping the uuid column, there's no need for it now
df_main.drop(columns=['uuid'], inplace=True)

# dictionary housing the sample type appreviations and all of their respective full names. if a new sample type is 
# ever introduced, just add another entry onto the dictionary.
dict_data_types = {
    'D.FLOW': 'Flow Cytometry',
    'CEL':'Cells',
    'LYS': 'Cell Extract',
    'BAC': 'Bacteria',
    'CEX': 'Cell Extract',
    'D.MSP': 'Mass Spectrometry',
    'A.MSP': 'Mass Spectrometry Analysis',
    'MUS': 'Mice',
    'TIS': 'Tissue',
    'D.IMG': 'Imaging',
    'D.SNSR': 'Sensor',
    'GAS': 'Gas',
    'D.ATM': 'Atmosphere',
    'DNA': 'DNA Library',
    'D.SEQ': 'Sequencing',
    'D.GPT': 'GPT Assay',
    'NHP': 'Non Human Primate',
    'D.LMX': 'Luminex',
    'A.IMG': 'Imaging Analysis', 
    'A.FLOW': 'Flow Cytometry Analysis',
    'A.LMX': 'Luminex Analysis',
    'D.FILE': 'File',
    'PAT': 'Human Patient', 
    'PAV': 'Patient Visit',
    'A.SCXP': 'Single Cell Expression Matrix Analysis', 
    'A.SCCL': 'Single Cell Clustering Analysis', 
    'LOC': 'Location', 
    'AB': 'Antibody', 
    'ABP': 'Antibody Panel',
    'CHM': 'Chemical', 
    'A.CTSC': 'CT Scan Analysis', 
    'OOC': 'Organ on Chip', 
    'D.FCS': 'Flow Cytometry', 
    'D.ELSA': 'Elisa', 
    'A.ALN': 'Genome Alignment Analysis', 
    'A.MUSP': 'Mutational Spectrum Analysis', 
    'D.CRY': 'Crystallography', 
    'D.PRM': 'Magnetic Resonance',
    'D.XRS': 'Spectroscopy',
    'D.NMR': 'Magnetic Resonance', 
    'A.NMR': 'Nuclear Magnetic Resonance Analysis', 
    'D.SPC': 'Spectroscopy', 
    'A.SPC': 'Spectrosctopy Analysis', 
    'RNA': 'RNA Library', 
    'A.CCE': 'Cancer Cell Extravasation Analysis', 
    'A.PERM': 'Permeability Analysis', 
    'A.CLOT': 'Clot Modeling Analysis', 
    'A.DBMM': 'Digitally Barcoded MTB Matrix Analysis'    
}

# replacing all the sample type abbreviated words in main dataframe with its full name
df_main['Sample Type'] = df_main['Sample Type'].replace(dict_data_types)

# empty_list, database_list, and fairdomhub_list are initialized as a 44 length array with zeros in it (doesn't export if all the arrays in the 
# dataframe are the same length). if another data type needs to be added add to the array + increase the number of zeros in the lists
empty_list = np.zeros(44)
data_list = ['Antibody', 'Bacteria', 'Cells', 'Chemical', 'Gas', 'Human Patient', 'Location', 'Mice', 'Non Human Primate', 'Organ on Chip', 'Patient Visit',
             'Antibody Panel','Cell Extract','DNA Library','RNA Library', 'Tissue',
             'Atmosphere','Crystallography', 'Elisa','File', 'Flow Cytometry','GPT Assay','Imaging','Luminex', 'Magnetic Resonance', 'Mass Spectrometry', 'Sensor','Sequencing','Spectroscopy',
             'Cancer Cell Extravasation Analysis','Clot Modeling Analysis','CT Scan Analysis','Digitally Barcoded MTB Matrix Analysis','Flow Cytometry Analysis','Genome Alignment Analysis','Imaging Analysis','Luminex Analysis','Mass Spectrometry Analysis','Mutational Spectrum Analysis','Nuclear Magnetic Resonance Analysis','Permeability Analysis','Single Cell Clustering Analysis','Single Cell Expression Matrix Analysis','Spectrosctopy Analysis']
database_list = np.zeros(44)
fairdomhub_list = np.zeros(44)

# one table for each project. these are NOT data frames, they are dictionaries. initialized into df later on in the code.
# key values in the dictionary are the column names for the final figure. the empty list is where classifications (source, raw) will later go
# impact table dictionary
table_impact = {
    '': empty_list,
    'Data Types': data_list,
    'MIT.SEEK': database_list,
    'Fairdomhub': fairdomhub_list}

# srp table dictionary
table_srp = {
    '': empty_list,
    'Data Types': data_list,
    'MIT.SEEK': database_list,
    'Fairdomhub': fairdomhub_list}

# metnet table dictionary
table_metnet = {
    '': empty_list,
    'Data Types': data_list,
    'MIT.SEEK': database_list,
    'Fairdomhub': fairdomhub_list}

# griffith table dictionary
table_griffith = {
    '': empty_list,
    'Data Types': data_list,
    'MIT.SEEK': database_list,
    'Fairdomhub': fairdomhub_list}

# four empty data frames initialized with their column names. classification column is not initialized – currently not needed 
df_impact = pd.DataFrame(columns=['Data Types', 'MIT.SEEK', 'Fairdomhub'])
df_srp = pd.DataFrame(columns=['Data Types', 'MIT.SEEK', 'Fairdomhub'])
df_metnet = pd.DataFrame(columns=['Data Types', 'MIT.SEEK', 'Fairdomhub'])
df_griffith = pd.DataFrame(columns=['Data Types', 'MIT.SEEK', 'Fairdomhub'])

# lab lists, used to check the row's lab names and associate them with the correct data frames initialized above
impact_labs = ['BRY', 'FLY', 'SED', 'NIH', 'NEM', 'ALT', 'BOO', 'LAL', 'SHA', 'LAU', 'FOR', 'BEH', 'SAS', 'SES']
srp_labs = ['HEM', 'ESS', 'KRO', 'SWA', 'WHI', 'SEL', 'WAD', 'ENG']
metnet_labs = ['KAM']
griffith_labs = ['GRI']

# start of filter code to filter the labs to their respective data frames
# initializing empty lists for each project. these will be used to house all dataframes associated with the project
list_df_impact = []
list_df_srp = []
list_df_metnet = []
list_df_griffith = []

# loops through each row of the main dataframe which currently has a 'sample type' and a 'lab' column. if the lab name matches the lab name
# of the project, the lab and sample type are initialized into a dataframe, which is added to the empty list initialized above.
# pd.concat (used later) works by iterating through a list of dataframes, why it's coded this way
for i in range(len(df_main)):
    # impact
    if df_main.loc[i, 'Lab'] in impact_labs:
        df_impact_concat = pd.DataFrame({'Data Types': [df_main.loc[i, 'Sample Type']],
                                  'MIT.SEEK': '',
                                  'Fairdomhub': ''})
        list_df_impact.append(df_impact_concat)
    # srp
    if df_main.loc[i, 'Lab'] in srp_labs:
        df_srp_concat = pd.DataFrame({'Data Types': [df_main.loc[i, 'Sample Type']],
                                  'MIT.SEEK': '',
                                  'Fairdomhub': ''})
        list_df_srp.append(df_srp_concat)
    # metnet
    if df_main.loc[i, 'Lab'] in metnet_labs:
        df_metnet_concat = pd.DataFrame({'Data Types': [df_main.loc[i, 'Sample Type']],
                                  'MIT.SEEK': '',
                                  'Fairdomhub': ''})
        list_df_metnet.append(df_metnet_concat)
    # griffith
    if df_main.loc[i, 'Lab'] in griffith_labs:
        df_griffith_concat = pd.DataFrame({'Data Types': [df_main.loc[i, 'Sample Type']],
                                  'MIT.SEEK': '',
                                  'Fairdomhub': ''})
        list_df_griffith.append(df_griffith_concat)

# pd.concat iterates through the list of data frames filled out above & adds them to the main dataframe. that dataframe now has 'sample type'
# and 'lab' information but only has the info for the labs that belong to that project
df_impact = pd.concat(list_df_impact, ignore_index=True)
df_srp = pd.concat(list_df_srp, ignore_index=True)
df_metnet = pd.concat(list_df_metnet, ignore_index=True)
df_griffith = pd.concat(list_df_griffith, ignore_index=True)

# make a new dataframe with the column names table_impact has (including the empty one that wasn't initialized in the previous df)
table_impact_df = pd.DataFrame(table_impact)
table_srp_df = pd.DataFrame(table_srp)
table_metnet_df = pd.DataFrame(table_metnet)
table_griffith_df = pd.DataFrame(table_griffith)

# function that takes a df with data, a length, and the new dataframe to be filled. if the data type is in the list of df + full name
# defined in the above dictionary, you get the index where that happens, get the current count at that index of the df, then add one to it.
def fill_in_table(df, length, table):
    for i in range(length):
        if df.loc[i, 'Data Types'] in table['Data Types'].values:
            index = table[table['Data Types'] == df.loc[i, 'Data Types']].index[0]
            count = int(table.at[index, 'MIT.SEEK'])
            table.at[index, 'MIT.SEEK'] = count + 1
    
# call the function for all four projects with the correct df and lengths
fill_in_table(df_impact, len(df_impact), table_impact_df)
fill_in_table(df_srp, len(df_srp), table_srp_df)
fill_in_table(df_metnet, len(df_metnet), table_metnet_df)
fill_in_table(df_griffith, len(df_griffith), table_griffith_df)

# drops any rows that has 0 as its value (meaning no samples are present for that data type in that project)
def drop_rows_with_zero(table_df, column_name):
    table_df.drop(table_df[table_df[column_name] == 0].index, inplace=True)
    table_df.reset_index(drop=True, inplace=True)

# called on all four dfs, check the column named 'MIT.SEEK' (where the counts are being incremented)
drop_rows_with_zero(table_impact_df, 'MIT.SEEK')
drop_rows_with_zero(table_srp_df, 'MIT.SEEK')
drop_rows_with_zero(table_metnet_df, 'MIT.SEEK')
drop_rows_with_zero(table_griffith_df, 'MIT.SEEK')

# after dropping all rows with 0, replace the 0s with 'None' so it looks cleaner when exported
def replace_zeros(table_df):
    table_df.replace(0, None, inplace=True)

# called on all four dfs
replace_zeros(table_impact_df)
replace_zeros(table_srp_df)
replace_zeros(table_metnet_df)
replace_zeros(table_griffith_df)

# separates all data types into four possible categories, this is purely for the category filter 
dict_data_types = {
    'Source Samples': ['Organ on Chip', 'Chemical', 'Antibody', 'Location', 'Patient Visit', 'Human Patient', 'Non Human Primate', 'Cells', 'Bacteria', 'Mice', 'Gas'],
    'Processed Samples': ['RNA Library', 'Antibody Panel', 'DNA Library', 'Cell Extract', 'Tissue', 'Cell Extract'],
    'Raw Data': ['Spectroscopy', 'Magnetic Resonance', 'Spectroscopy', 'Magnetic Resonance', 'Crystallography', 'Elisa', 'Flow Cytometry', 'File', 'Luminex', 'GPT Assay', 'Sequencing', 'Atmosphere', 'Flow Cytometry', 'Mass Spectrometry', 'Imaging', 'Sensor'],
    'Analyzed Data': ['Digitally Barcoded MTB Matrix Analysis' , 'Clot Modeling Analysis', 'Permeability Analysis', 'Cancer Cell Extravasation Analysis', 'Spectrosctopy Analysis', 'Nuclear Magnetic Resonance Analysis', 'Mutational Spectrum Analysis', 'Genome Alignment Analysis', 'CT Scan Analysis', 'Single Cell Clustering Analysis', 'Single Cell Expression Matrix Analysis', 'Luminex Analysis', 'Flow Cytometry Analysis', 'Imaging Analysis', 'Mass Spectrometry Analysis']}

# function that takes the data types in each dataframe and replaces it with 'source samples', 'processed samples' 'raw data', or 'analyzed data'
def map_data_to_category(data_type):
    for category, data_list in dict_data_types.items():
        if data_type in data_list:
            return category
    return None

# apply the mapping function to 'Data Types' column, changes all the data types to their categories
table_impact_df[''] = table_impact_df['Data Types'].apply(map_data_to_category)
table_srp_df[''] = table_srp_df['Data Types'].apply(map_data_to_category)
table_metnet_df[''] = table_metnet_df['Data Types'].apply(map_data_to_category)
table_griffith_df[''] = table_griffith_df['Data Types'].apply(map_data_to_category)

# bunch of funtions here that deal with style + visual components of the final graph. some of these are reused for the lab specific figures.

# applies a color to every row, takes in row and column values
def apply_color_to_rows(sheet, row_start, row_end, col_start, col_end, color):
    for row in sheet.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end):
        for cell in row:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
# applies a black bottom border to each cell.
def apply_bottom_border(sheet, col_start, col_end):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_start, max_col=col_end):
        apply_bottom_border_to_row(sheet, row)

# helper function for above
def apply_bottom_border_to_row(sheet, row):
    for cell in row:
        cell.border = Border(bottom=Side(border_style='thin'))
        
# autosize function, autosizes the width of the columns based on how long the string inside that cell is. has a 'max' function so that you 
# can get the maximum size of the longest string and autosize it to that
def autosize(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
        
# sets a hard coded length for the height of the row, not based on text size or anything
def set_row_height(worksheet, row_height_in_points):
    for row in worksheet.iter_rows():
        for cell in row:
            worksheet.row_dimensions[cell.row].height = row_height_in_points
         
# sets a row height but for a specific row index in the figure
def set_row_height_idx(worksheet, row_index, row_height_in_points):
    worksheet.row_dimensions[row_index].height = row_height_in_points
        
# bolds the first column in the figure (for the headers)
def bold_column(sheet):
    for cell in sheet['A']:
        cell.font = Font(bold=True)

# bolds the first row in the figure (for the headers)
def bold_row(sheet):
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        
# center aligns data, takes in row and column values
def center_data(min_row, max_row, min_col, max_col, sheet):
    for row in sheet.iter_rows(min_row, max_row, min_col, max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

# right aligns data, takes in row and column values
def right_data(min_row, max_row, min_col, max_col, sheet):
    for row in sheet.iter_rows(min_row, max_row, min_col, max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="center")

# center aligns data vertically, takes in row and column values. only used for metnet/griffith because those have rows where the height is larger
# to accomodate for the category name
def center_data_vert(min_row, max_row, min_col, max_col, sheet):
    for row in sheet.iter_rows(min_row, max_row, min_col, max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

# strips of any leading/trailing whitespace
table_impact_df[''] = table_impact_df[''].str.strip()

# grabs the first and last instances of each category and grabs the index of two. +2 is added to account for excels' header taking up one
# row and figure header taking up one row. these indices are necessary to figure out what rows to merge together for category header
first_idx_ss_ip = table_impact_df[table_impact_df[''] == 'Source Samples'].index[0] + 2
last_idx_ss_ip = table_impact_df[table_impact_df[''] == 'Source Samples'].index[-1] + 2
first_idx_ps_ip = table_impact_df[table_impact_df[''] == 'Processed Samples'].index[0] + 2
last_idx_ps_ip = table_impact_df[table_impact_df[''] == 'Processed Samples'].index[-1] + 2
first_idx_rd_ip = table_impact_df[table_impact_df[''] == 'Raw Data'].index[0] + 2
last_idx_rd_ip = table_impact_df[table_impact_df[''] == 'Raw Data'].index[-1] + 2
first_idx_ad_ip = table_impact_df[table_impact_df[''] == 'Analyzed Data'].index[0] + 2
last_idx_ad_ip = table_impact_df[table_impact_df[''] == 'Analyzed Data'].index[-1] + 2

# this replaces all except the first row of the to-be-merged category rows with whitspace. this leaves the first row of each category
# with a classification and the rest blank until we get to the next classification
table_impact_df.loc[first_idx_ss_ip - 1:last_idx_ss_ip - 2, ''] = ''
table_impact_df.loc[first_idx_ps_ip - 1:last_idx_ps_ip - 2, ''] = ''
table_impact_df.loc[first_idx_rd_ip - 1:last_idx_rd_ip - 2, ''] = ''
table_impact_df.loc[first_idx_ad_ip - 1:last_idx_ad_ip - 2, ''] = ''

# same code except for the srp project
table_srp_df[''] = table_srp_df[''].str.strip()

first_idx_ss_sp = table_srp_df[table_srp_df[''] == 'Source Samples'].index[0] + 2
last_idx_ss_sp = table_srp_df[table_srp_df[''] == 'Source Samples'].index[-1] + 2
first_idx_ps_sp = table_srp_df[table_srp_df[''] == 'Processed Samples'].index[0] + 2
last_idx_ps_sp = table_srp_df[table_srp_df[''] == 'Processed Samples'].index[-1] + 2
first_idx_rd_sp = table_srp_df[table_srp_df[''] == 'Raw Data'].index[0] + 2
last_idx_rd_sp = table_srp_df[table_srp_df[''] == 'Raw Data'].index[-1] + 2
first_idx_ad_sp = table_srp_df[table_srp_df[''] == 'Analyzed Data'].index[0] + 2
last_idx_ad_sp = table_srp_df[table_srp_df[''] == 'Analyzed Data'].index[-1] + 2

table_srp_df.loc[first_idx_ss_sp - 1:last_idx_ss_sp - 2, ''] = ''
table_srp_df.loc[first_idx_ps_sp - 1:last_idx_ps_sp - 2, ''] = ''
table_srp_df.loc[first_idx_rd_sp - 1:last_idx_rd_sp - 2, ''] = ''
table_srp_df.loc[first_idx_ad_sp - 1:last_idx_ad_sp - 2, ''] = ''

# griffith
table_griffith_df[''] = table_griffith_df[''].str.strip()

first_idx_ss_gr = table_griffith_df[table_griffith_df[''] == 'Source Samples'].index[0] + 2
last_idx_ss_gr = table_griffith_df[table_griffith_df[''] == 'Source Samples'].index[-1] + 2
first_idx_ps_gr = table_griffith_df[table_griffith_df[''] == 'Processed Samples'].index[0] + 2
last_idx_ps_gr = table_griffith_df[table_griffith_df[''] == 'Processed Samples'].index[-1] + 2
first_idx_rd_gr = table_griffith_df[table_griffith_df[''] == 'Raw Data'].index[0] + 2
last_idx_rd_gr = table_griffith_df[table_griffith_df[''] == 'Raw Data'].index[-1] + 2
first_idx_ad_gr = table_griffith_df[table_griffith_df[''] == 'Analyzed Data'].index[0] + 2
last_idx_ad_gr = table_griffith_df[table_griffith_df[''] == 'Analyzed Data'].index[-1] + 2

table_griffith_df.loc[first_idx_ss_gr - 1:last_idx_ss_gr - 2, ''] = ''
table_griffith_df.loc[first_idx_ps_gr - 1:last_idx_ps_gr - 2, ''] = ''
table_griffith_df.loc[first_idx_rd_gr - 1:last_idx_rd_gr - 2, ''] = ''
table_griffith_df.loc[first_idx_ad_gr - 1:last_idx_ad_gr - 2, ''] = ''

# metnet
table_metnet_df[''] = table_metnet_df[''].str.strip()

first_idx_ss_mn = table_metnet_df[table_metnet_df[''] == 'Source Samples'].index[0] + 2
last_idx_ss_mn = table_metnet_df[table_metnet_df[''] == 'Source Samples'].index[-1] + 2
first_idx_ps_mn = table_metnet_df[table_metnet_df[''] == 'Processed Samples'].index[0] + 2
last_idx_ps_mn = table_metnet_df[table_metnet_df[''] == 'Processed Samples'].index[-1] + 2
first_idx_rd_mn = table_metnet_df[table_metnet_df[''] == 'Raw Data'].index[0] + 2
last_idx_rd_mn = table_metnet_df[table_metnet_df[''] == 'Raw Data'].index[-1] + 2
first_idx_ad_mn = table_metnet_df[table_metnet_df[''] == 'Analyzed Data'].index[0] + 2
last_idx_ad_mn = table_metnet_df[table_metnet_df[''] == 'Analyzed Data'].index[-1] + 2

table_srp_df.loc[first_idx_ss_mn - 1:last_idx_ss_mn - 2, ''] = ''
table_srp_df.loc[first_idx_ps_mn - 1:last_idx_ps_mn - 2, ''] = ''
table_srp_df.loc[first_idx_rd_mn - 1:last_idx_rd_mn - 2, ''] = ''
table_srp_df.loc[first_idx_ad_mn - 1:last_idx_ad_mn - 2, ''] = ''

# load the existing excel file + turn it into a book
existing_file = 'Figure Generation Samples.xlsx'
book = load_workbook(existing_file)

# merges + rotates, helper function to be called on each of the four projects with their respective indices. also sets text rotation
# + centers as a default. wraptext ensures the text doesn't flow onto other cells
def merge_rotate(sheet, first, last, color=None):
    sheet.merge_cells('A' + str(first) + ':A' + str(last))
    cell = sheet['A' + str(first)]
    cell.alignment = Alignment(textRotation=90, wrapText=True, horizontal="center", vertical="center")

# impact
# make a new sheet
new_sheet_impact = 'Impact Figure'

# make a new sheet onto the existing file
new_sheet_created_impact = book.create_sheet(title=new_sheet_impact)

# load all rows of the df carrying all the data into the new sheet
for r in dataframe_to_rows(table_impact_df, index=False, header=True):
    new_sheet_created_impact.append(r)

# moving the data into center or right depending on what it is. this is purposely run before the merge & rotate because
# if you run it after the text rotation gets overpowered + doesn't show up by the new alignment
right_data(2, new_sheet_created_impact.max_row, 2, 2, new_sheet_created_impact)
center_data(1, 1, 2, 4, new_sheet_created_impact)

# merge + rotate each one based on the index of the source sample, processed sample, raw data, and analyzed data
merge_rotate(new_sheet_created_impact, first_idx_ss_ip, last_idx_ss_ip) 
merge_rotate(new_sheet_created_impact, first_idx_ps_ip, last_idx_ps_ip) 
merge_rotate(new_sheet_created_impact, first_idx_rd_ip, last_idx_rd_ip) 
merge_rotate(new_sheet_created_impact, first_idx_ad_ip, last_idx_ad_ip) 

# adding a bottom border to all rows + specified columns
apply_bottom_border(new_sheet_created_impact, col_start=2, col_end=5)

# adding color to the rows based on the category indices
apply_color_to_rows(new_sheet_created_impact, row_start=first_idx_ss_ip, row_end=last_idx_ss_ip, col_start=1, col_end=5, color="74C476")
apply_color_to_rows(new_sheet_created_impact, row_start=first_idx_ps_ip, row_end=last_idx_ps_ip, col_start=1, col_end=5, color="FE9929")
apply_color_to_rows(new_sheet_created_impact, row_start=first_idx_rd_ip, row_end=last_idx_rd_ip, col_start=1, col_end=5, color="CFDEF2")
apply_color_to_rows(new_sheet_created_impact, row_start=first_idx_ad_ip, row_end=last_idx_ad_ip, col_start=1, col_end=5, color="5496D2")

# bolding the first column + first row
bold_column(new_sheet_created_impact)
bold_row(new_sheet_created_impact)

# setting the row height to a bit bigger so it doens't look so squashed
set_row_height(new_sheet_created_impact, row_height_in_points=17)

# autorsizes width of cells
autosize(new_sheet_created_impact)

# the same code but for srp
new_sheet_srp = 'SRP Figure'
new_sheet_created_srp = book.create_sheet(title=new_sheet_srp)

for r in dataframe_to_rows(table_srp_df, index=False, header=True):
    new_sheet_created_srp.append(r)

right_data(2, new_sheet_created_srp.max_row, 2, 2, new_sheet_created_srp)
center_data(1, 1, 2, 4, new_sheet_created_srp)

merge_rotate(new_sheet_created_srp, first_idx_ss_sp, last_idx_ss_sp)
merge_rotate(new_sheet_created_srp, first_idx_ps_sp, last_idx_ps_sp)
merge_rotate(new_sheet_created_srp, first_idx_rd_sp, last_idx_rd_sp)
merge_rotate(new_sheet_created_srp, first_idx_ad_sp, last_idx_ad_sp)

apply_bottom_border(new_sheet_created_srp, col_start=2, col_end=5)
apply_color_to_rows(new_sheet_created_srp, row_start=first_idx_ss_sp, row_end=last_idx_ss_sp, col_start=1, col_end=5, color="74C476")
apply_color_to_rows(new_sheet_created_srp, row_start=first_idx_ps_sp, row_end=last_idx_ps_sp, col_start=1, col_end=5, color="FE9929")
apply_color_to_rows(new_sheet_created_srp, row_start=first_idx_rd_sp, row_end=last_idx_rd_sp, col_start=1, col_end=5, color="CFDEF2")
apply_color_to_rows(new_sheet_created_srp, row_start=first_idx_ad_sp, row_end=last_idx_ad_sp, col_start=1, col_end=5, color="5496D2")

bold_column(new_sheet_created_srp)
bold_row(new_sheet_created_srp)

set_row_height(new_sheet_created_srp, row_height_in_points=17)

autosize(new_sheet_created_srp)

# metnet
new_sheet_metnet = 'MetNet Figure'
new_sheet_created_metnet = book.create_sheet(title=new_sheet_metnet)

for r in dataframe_to_rows(table_metnet_df, index=False, header=True):
    new_sheet_created_metnet.append(r)

right_data(2, new_sheet_created_metnet.max_row, 2, 2, new_sheet_created_metnet)
center_data(1, 1, 2, 4, new_sheet_created_metnet)

merge_rotate(new_sheet_created_metnet, first_idx_ss_mn, last_idx_ss_mn)
merge_rotate(new_sheet_created_metnet, first_idx_ps_mn, last_idx_ps_mn)
merge_rotate(new_sheet_created_metnet, first_idx_rd_mn, last_idx_rd_mn)
merge_rotate(new_sheet_created_metnet, first_idx_ad_mn, last_idx_ad_mn)

apply_bottom_border(new_sheet_created_metnet, col_start=2, col_end=5)
apply_color_to_rows(new_sheet_created_metnet, row_start=first_idx_ss_mn, row_end=last_idx_ss_mn, col_start=1, col_end=5, color="74C476")
apply_color_to_rows(new_sheet_created_metnet, row_start=first_idx_ps_mn, row_end=last_idx_ps_mn, col_start=1, col_end=5, color="FE9929")
apply_color_to_rows(new_sheet_created_metnet, row_start=first_idx_rd_mn, row_end=last_idx_rd_mn, col_start=1, col_end=5, color="CFDEF2")
apply_color_to_rows(new_sheet_created_metnet, row_start=first_idx_ad_mn, row_end=last_idx_ad_mn, col_start=1, col_end=5, color="5496D2")

set_row_height(new_sheet_created_metnet, row_height_in_points=17)

# this sets the row higher to make room for the entire category name
set_row_height_idx(new_sheet_created_metnet, row_index=5, row_height_in_points=58)

bold_column(new_sheet_created_metnet)
bold_row(new_sheet_created_metnet)

# this bit is to center the data in the thicker rows to make sure the data is centered vertically
center_data_vert(2, new_sheet_created_metnet.max_row, 3, 3, new_sheet_created_metnet)

autosize(new_sheet_created_metnet)

# griffith
new_sheet_griffith = 'Griffith Figure'
new_sheet_created_griffith = book.create_sheet(title=new_sheet_griffith)

for r in dataframe_to_rows(table_griffith_df, index=False, header=True):
    new_sheet_created_griffith.append(r)

right_data(2, new_sheet_created_griffith.max_row, 2, 2, new_sheet_created_griffith)
center_data(1, 1, 2, 4, new_sheet_created_griffith)

merge_rotate(new_sheet_created_griffith, first_idx_ss_gr, last_idx_ss_gr)
merge_rotate(new_sheet_created_griffith, first_idx_ps_gr, last_idx_ps_gr)
merge_rotate(new_sheet_created_griffith, first_idx_rd_gr, last_idx_rd_gr)
merge_rotate(new_sheet_created_griffith, first_idx_ad_gr, last_idx_ad_gr)

apply_bottom_border(new_sheet_created_griffith, col_start=2, col_end=5)
apply_color_to_rows(new_sheet_created_griffith, row_start=first_idx_ss_gr, row_end=last_idx_ss_gr, col_start=1, col_end=5, color="74C476")
apply_color_to_rows(new_sheet_created_griffith, row_start=first_idx_ps_gr, row_end=last_idx_ps_gr, col_start=1, col_end=5, color="FE9929")
apply_color_to_rows(new_sheet_created_griffith, row_start=first_idx_rd_gr, row_end=last_idx_rd_gr, col_start=1, col_end=5, color="CFDEF2")
apply_color_to_rows(new_sheet_created_griffith, row_start=first_idx_ad_gr, row_end=last_idx_ad_gr, col_start=1, col_end=5, color="5496D2")

set_row_height(new_sheet_created_griffith, row_height_in_points=17)
set_row_height_idx(new_sheet_created_griffith, row_index=9, row_height_in_points=56)

# there's two samples for processed, changing both the row heights so it looks more uniform
set_row_height_idx(new_sheet_created_griffith, row_index=5, row_height_in_points=30)
set_row_height_idx(new_sheet_created_griffith, row_index=6, row_height_in_points=30)

bold_column(new_sheet_created_griffith)
bold_row(new_sheet_created_griffith)

center_data_vert(2, new_sheet_created_griffith.max_row, 3, 3, new_sheet_created_griffith)

autosize(new_sheet_created_griffith)

#### LAB FIGURES 

# df_main_lab = all the data the script is going to read in. same deal as above, just named something differently so there's a clear differentiation
# from here on it's mostly the same deal as above but with different filling in techniques.
file_path_lab = 'Figure Generation Samples.xlsx'
sheet_name_lab = 'Data'
df_main_lab = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=1)
df_main_lab.reset_index(drop=True, inplace=True)

# making sample type and lab column
df_main_lab['Sample Type'] = df_main_lab['uuid'].apply(get_sample_type)
df_main_lab['Lab'] = df_main_lab['uuid'].apply(get_lab)

df_main_lab.drop(columns=['uuid'], inplace=True)

# hard coded in lab names based on the order which it looks best, can play around with this when new samples come in
lab_list_impact = ['NIH', 'NEM', 'SED', 'LAL', 'SAS', 'ALT', 'BOO', 'FLY', 'SES', 'FOR', 'SHA', 'BEH', 'BRY', 'LAU']
lab_list_srp = ['ENG', 'ESS', 'WHI', 'SWA', 'HEM', 'KRO', 'SEL', 'WAD']
lab_list_metnet = ['KAM']
lab_list_griffith = ['GRI']

# categories filled in with their abbreviated data type names
dict_data_types = {
    'Source Samples': ['OOC', 'CHM', 'AB', 'LOC', 'PAV', 'PAT', 'NHP', 'CEL', 'BAC', 'MUS', 'GAS'],
    'Processed Samples': ['RNA', 'ABP', 'DNA', 'LYS', 'TIS', 'CEX'],
    'Raw Data': ['D.SPC', 'D.NMR', 'D.XRS', 'D.PRM', 'D.CRY', 'D.ELSA', 'D.FCS', 'D.FILE', 'D.LMX', 'D.GPT', 'D.SEQ', 'D.ATM', 'D.FLOW', 'D.MSP', 'D.IMG', 'D.SNSR'],
    'Analyzed Data': ['A.DBMM', 'A.CLOT', 'A.PERM', 'A.CCE', 'A.SPC', 'A.NMR', 'A.MUSP', 'A.ALN', 'A.CTSC', 'A.SCCL', 'A.SCXP', 'A.LMX', 'A.FLOW', 'A.IMG', 'A.MSP']}

# replaces the abbreviated data type names with 'source samples', 'processed samples', 'raw data', or 'analyzed data'
df_main_lab['Sample Type'] = df_main_lab['Sample Type'].replace({val: key for key, values in dict_data_types.items() for val in values})

# one table for each project. these ARE data frames. same deal as above: np.zeroes there to ensure same length, and each one has its lab list
table_impact_lab = pd.DataFrame({
    '': lab_list_impact,
    'Source Samples': np.zeros(len(lab_list_impact)),
    'Processed Samples': np.zeros(len(lab_list_impact)),
    'Raw Data': np.zeros(len(lab_list_impact)),
    'Analyzed Data': np.zeros(len(lab_list_impact))
})

# srp
table_srp_lab = pd.DataFrame({
    '': lab_list_srp,
    'Source Samples': np.zeros(len(lab_list_srp)),
    'Processed Samples': np.zeros(len(lab_list_srp)),
    'Raw Data': np.zeros(len(lab_list_srp)),
    'Analyzed Data': np.zeros(len(lab_list_srp))
})

# metnet
table_metnet_lab = pd.DataFrame({
    '': lab_list_metnet,
    'Source Samples': np.zeros(len(lab_list_metnet)),
    'Processed Samples': np.zeros(len(lab_list_metnet)),
    'Raw Data': np.zeros(len(lab_list_metnet)),
    'Analyzed Data': np.zeros(len(lab_list_metnet))
})

# griffith
table_griffith_lab = pd.DataFrame({
    '': lab_list_griffith,
    'Source Samples': np.zeros(len(lab_list_griffith)),
    'Processed Samples': np.zeros(len(lab_list_griffith)),
    'Raw Data': np.zeros(len(lab_list_griffith)),
    'Analyzed Data': np.zeros(len(lab_list_griffith))
})

# filling in the individual tables. for each row, you get the lab and the sample type, and update with the corresponding lab name and df.
# updates the count at the very end adding to the one already there
for i in range(len(df_main_lab)):
    lab = df_main_lab.loc[i, 'Lab']
    sample_type = df_main_lab.loc[i, 'Sample Type'].strip()

    if lab in lab_list_impact:
        table = table_impact_lab
    elif lab in lab_list_srp:
        table = table_srp_lab
    elif lab in lab_list_metnet:
        table = table_metnet_lab
    elif lab in lab_list_griffith:
        table = table_griffith_lab
    else:
        continue  # Skip rows with unknown labs

    # Find the row(s) that match the lab
    rows_to_update = table[''] == lab

    # Make a copy of the subset to avoid modifying the original DataFrame
    table_subset = table.loc[rows_to_update, sample_type].copy()

    # Update the counts
    table.loc[rows_to_update, sample_type] = table_subset + 1

    # Replacing lab names with their actual names
# impact
impact_map_lab = {
    'BRY': 'Bryson',
    'FLY': 'Flynn',
    'SED': 'Seder',
    'NIH': 'NIH',
    'NEM': 'Nemes',
    'ALT': 'Alter',
    'BOO': 'Boom',
    'LAL': 'Lalvani',
    'SHA': 'Shalek',
    'LAU': 'Lauffenburger',
    'FOR': 'Fortune',
    'BEH': 'Behar',
    'SAS': 'Sassetti',
    'SES': 'Seshadri'
}

# srp
srp_map_lab = {'HEM': 'Hemond', 
           'ESS': 'Essigmann', 
           'KRO': 'Keoll', 
           'SWA': 'Swager', 
           'WHI': 'White', 
           'SEL': 'Selin', 
           'WAD': 'Wadduwage', 
           'ENG': 'Engelward'
}

# metnet
metnet_map_lab = {
    'KAM': 'Kamm'
}

# griffith
griffith_map_lab = {
    'GRI' : 'Griffith'
}

# replacing all lab names in each of the df with their full name. these have an empty as a column name, the lab list is 
# going to be the row name
table_impact_lab[''] = table_impact_lab[''].map(impact_map_lab).fillna(table_impact_lab[''])
table_srp_lab[''] = table_srp_lab[''].map(srp_map_lab).fillna(table_srp_lab[''])
table_metnet_lab[''] = table_metnet_lab[''].map(metnet_map_lab).fillna(table_metnet_lab[''])
table_griffith_lab[''] = table_griffith_lab[''].map(griffith_map_lab).fillna(table_griffith_lab[''])

# creates the sum row for impact. takes the sum of each column above it and places it in a new dataframe
sum_row_impact = pd.DataFrame({'': '',
                               'Source Samples': [table_impact_lab['Source Samples'].sum()],
                               'Processed Samples': [table_impact_lab['Processed Samples'].sum()],
                               'Raw Data': [table_impact_lab['Raw Data'].sum()],
                               'Analyzed Data': [table_impact_lab['Analyzed Data'].sum()]
                              })

# that new dataframe is added onto the existing dataframe
table_impact_lab = pd.concat([table_impact_lab, sum_row_impact], ignore_index=True)

# same deal for srp
sum_row_srp = pd.DataFrame({'': '',
                               'Source Samples': [table_srp_lab['Source Samples'].sum()],
                               'Processed Samples': [table_srp_lab['Processed Samples'].sum()],
                               'Raw Data': [table_srp_lab['Raw Data'].sum()],
                               'Analyzed Data': [table_srp_lab['Analyzed Data'].sum()]
                              })

table_srp_lab = pd.concat([table_srp_lab, sum_row_srp], ignore_index=True)

# griffith + metnet both have one lab, don't need a sum row for that. if needed in the future take the code above and change
# the project and lab name

# replacing zeroes
replace_zeros(table_impact_lab)
replace_zeros(table_srp_lab)
replace_zeros(table_metnet_lab)
replace_zeros(table_griffith_lab)

# coloring in, doesn't take in row and color values but just applies color if the cell is filled with a numerical value
def apply_color(sheet, color):
    total_rows = sheet.max_row
    for row_num, row in enumerate(sheet.iter_rows(), start=1):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
# same code but doesn't fill in color for the last row, used with the srp + impact beacuse we don't wnt the sum row to have colored
# in values where the numerical values are
def apply_color_no_last_row(sheet, color):
    total_rows = sheet.max_row
    for row_num, row in enumerate(sheet.iter_rows(), start=1):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)):
                # Check if the current row is not the last row
                if row_num != total_rows:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
# same deal as above: making a new sheet, applying styling to it
# impact
new_sheet_impact_lab = 'Impact Figure Lab'
new_sheet_created_impact_lab = book.create_sheet(title=new_sheet_impact_lab)

for r in dataframe_to_rows(table_impact_lab, index=False, header=True):
    new_sheet_created_impact_lab.append(r)

apply_color_no_last_row(new_sheet_created_impact_lab, color="ffd966")
autosize(new_sheet_created_impact_lab)

bold_column(new_sheet_created_impact_lab)
bold_row(new_sheet_created_impact_lab)

set_row_height(new_sheet_created_impact_lab, row_height_in_points=17)

center_data(1, 1, 2, 5, new_sheet_created_impact_lab)
right_data(2, new_sheet_created_impact_lab.max_row, 1, 1, new_sheet_created_impact_lab)
center_data(2, new_sheet_created_impact_lab.max_row, 2, 5, new_sheet_created_impact_lab)

# srp
new_sheet_srp_lab = 'SRP Figure Lab'
new_sheet_created_srp_lab = book.create_sheet(title=new_sheet_srp_lab)

for r in dataframe_to_rows(table_srp_lab, index=False, header=True):
    new_sheet_created_srp_lab.append(r)

apply_color_no_last_row(new_sheet_created_srp_lab, color="ffd966")
autosize(new_sheet_created_srp_lab)

bold_column(new_sheet_created_srp_lab)
bold_row(new_sheet_created_srp_lab)

set_row_height(new_sheet_created_srp_lab, row_height_in_points=17)

center_data(1, 1, 2, 5, new_sheet_created_srp_lab)
right_data(2, new_sheet_created_srp_lab.max_row, 1, 1, new_sheet_created_srp_lab)
center_data(2, new_sheet_created_srp_lab.max_row, 2, 5, new_sheet_created_srp_lab)

# metnet
new_sheet_metnet_lab = 'MetNet Figure Lab'
new_sheet_created_metnet_lab = book.create_sheet(title=new_sheet_metnet_lab)

for r in dataframe_to_rows(table_metnet_lab, index=False, header=True):
    new_sheet_created_metnet_lab.append(r)
    
apply_color(new_sheet_created_metnet_lab, color="ffd966")
autosize(new_sheet_created_metnet_lab)

bold_column(new_sheet_created_metnet_lab)
bold_row(new_sheet_created_metnet_lab)

set_row_height(new_sheet_created_metnet_lab, row_height_in_points=17)

center_data(1, 1, 2, 5, new_sheet_created_metnet_lab)
right_data(2, new_sheet_created_metnet_lab.max_row, 1, 1, new_sheet_created_metnet_lab)
center_data(2, new_sheet_created_metnet_lab.max_row, 2, 5, new_sheet_created_metnet_lab)
    
# griffith
new_sheet_griffith_lab = 'Griffith Figure Lab'
new_sheet_created_griffith_lab = book.create_sheet(title=new_sheet_griffith_lab)

for r in dataframe_to_rows(table_griffith_lab, index=False, header=True):
    new_sheet_created_griffith_lab.append(r)

apply_color(new_sheet_created_griffith_lab, color="ffd966")
autosize(new_sheet_created_griffith_lab)

bold_column(new_sheet_created_griffith_lab)
bold_row(new_sheet_created_griffith_lab)

set_row_height(new_sheet_created_griffith_lab, row_height_in_points=17)

center_data(1, 1, 2, 5, new_sheet_created_griffith_lab)
right_data(2, new_sheet_created_griffith_lab.max_row, 1, 1, new_sheet_created_griffith_lab)
center_data(2, new_sheet_created_griffith_lab.max_row, 2, 5, new_sheet_created_griffith_lab)

# saving all new sheets to the book + exporting to the same excel file
book.save(existing_file)

# prints 'done' when done
print("done")
